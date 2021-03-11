'''
휴대폰과 통신사 무제한 문자로 그룹문자 보내는 파이썬 코드
https://godpeople.or.kr/board/3417787

핸드폰의 무료 문자를 이용해서 단체문자를 쉽게 보낼 수 있는 파이썬 프로그램입니다.

먼저, 핸드폰에 AirMore라는 앱을 설치합니다.
Install AirMore App and Download

안드로이드
https://play.google.com/store/apps/details?id=com.airmore
아이폰
https://itunes.apple.com/us/app/airmore/id997212086?ls=1&mt=8

그리고 공유기에 연결된 핸드폰의 아이피를 확인합니다.
컴퓨터에 첨부파일을 다운받은 후, message.txt 파일과 contact.xlsx 파일을 자신에 맞게 수정합니다.
실행시 아이피를 입력하면 보낼 사람과 번호가 나오게 됩니다.
문제가 없으면 발송하시면 됩니다.
파이썬 코드는 공개되어 있으며, 프로그램은 첨부파일로 받으시기 바랍니다.


1. 안드로이드 핸드폰에 airmore 설치
2. airmore 에서 와이파이 주소 확인
3. 파이썬 코드 자신에게 맞게 수정후 실행

최초 실행시, airmore에서 접속 확인을 해주어야 함.

테스트 후기.
1. 한글 깨짐없음.
2. mms 가능.
3. 그룹메시지 가능.
4. 보안 안정적
5. 사진 첨부파일이 불가능함. ㅠ
다른 방식이 있는지는 확인해봐야하는데, 일단 비용이 무료이고 파이썬 작동으로 상당히 안정적인 발송이 가능함. 해당 코드를 gui를 입혀서 프로그램으로 만들어도 좋을 것 같다는 생각임.

활용하기 좋은 점.
매일, 혹은 매주 정해진 시간대에 자동으로 문자를 발송하는 자동화 프로그램이 가능함.

안드로이드
https://play.google.com/store/apps/details?id=com.airmore
아이폰
https://itunes.apple.com/us/app/airmore/id997212086?ls=1&mt=8
각 통신사 무제한 문자 정책
https://godpeople.or.kr/board/3309836
'''

import pyairmore
from openpyxl import load_workbook
from ipaddress import IPv4Address  # for your IP address
from pyairmore.request import AirmoreSession  # to create an AirmoreSession
from pyairmore.services.messaging import MessagingService  # to send messages

# Airmore 활성화
ipAddress = input("핸드폰과 연결할 내부아이피를 입력하세요. ")
if not ipAddress:
    ip = IPv4Address("192.168.100.210")
else:
    ip = IPv4Address(ipAddress)

session = AirmoreSession(ip, 2333)
service = MessagingService(session)

# 보낼 문자내용
filepathMessage = open("./message.txt", 'rt', encoding='UTF8')
message = filepathMessage.read()
filepathMessage.close()

# 주소록 파일
filepathContact = "./contact.xlsx"

# 연락처 컬럼
columnName = "A"
columnPhone = "B"

workbook = load_workbook(filename=filepathContact, read_only=True)
worksheet = workbook.worksheets[0]  # 첫번째 시트를 찾음

# 행과 열의 갯수를 찾음
row_count = worksheet.max_row
column_count = worksheet.max_column

phoneNumbers = {}

for i in range(row_count):
    cellPhone = "{}{}".format(columnPhone, i + 1)
    cellName = "{}{}".format(columnName, i + 1)
    name = worksheet[cellName].value
    number = worksheet[cellPhone].value
    if number != "" or number is not None:
        phoneNumbers[name] = str(number)

for nm, pn in phoneNumbers.items():
    sendMessage = "{name}님 안녕하세요. {message}".format(name=nm, message=message)
    print(nm)
    print(pn)
    print(sendMessage)
    print("")

print("총 {count}명에게 문자메시지 발송 가능합니다.\n".format(count=len(phoneNumbers)))

order = input("위의 연락처로 문자를 발송하시겠습니까? Y or N ")

if order == "Y" or order == "YES" or order == "y" or order == "yes" or order == "Yes":
    for nm, pn in phoneNumbers.items():
        sendMessage = "{name}님 안녕하세요. {message}".format(name=nm, message=message)
        service.send_message(pn, sendMessage)
        print("{name}님의 {number}로 문자를 발송하였습니다.".format(name=nm, number=pn))
else:
    print("발송이 취소되었습니다.")

input("아무 키나 누르면 종료합니다. ")

